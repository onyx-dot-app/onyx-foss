"""Tests for the parsed spreadsheet preview path used by
GET /chat/file/{file_id}?parsed=true (xlsx chat attachments rendered as tables
in the frontend instead of raw binary bytes)."""

import csv
import io
from typing import cast
from unittest.mock import patch

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from onyx.configs.constants import FileOrigin
from onyx.file_processing.file_types import SPREADSHEET_MIME_TYPE
from onyx.file_store.file_store import get_default_file_store
from onyx.file_store.models import ChatFileType
from onyx.server.query_and_chat.chat_utils import (
    is_spreadsheet_mime_type,
    mime_type_to_chat_file_type,
    parse_spreadsheet_for_preview,
)


def _build_xlsx() -> bytes:
    workbook = openpyxl.Workbook()

    first_sheet = cast(Worksheet, workbook.active)
    first_sheet.title = "Revenue"
    first_sheet.append(["Region", "Quarter", "Amount"])
    first_sheet.append(["EMEA", "Q1", 1250.5])
    # Value with a comma + quote to exercise CSV escaping
    first_sheet.append(['APAC, "South"', "Q2", 900])

    second_sheet = workbook.create_sheet("Notes")
    second_sheet.append(["Note"])
    second_sheet.append(["multi\nline"])

    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


@pytest.mark.usefixtures("db_session", "tenant_context", "initialize_file_store")
def test_parse_spreadsheet_for_preview_from_file_store() -> None:
    """Round-trip: save an xlsx to the real file store, read it back the same
    way the endpoint does, and verify the parsed per-sheet CSV payload."""
    file_store = get_default_file_store()
    file_id = file_store.save_file(
        content=io.BytesIO(_build_xlsx()),
        display_name="report.xlsx",
        file_origin=FileOrigin.CHAT_UPLOAD,
        file_type=SPREADSHEET_MIME_TYPE,
    )

    file_record = file_store.read_file_record(file_id)
    assert file_record is not None
    assert is_spreadsheet_mime_type(file_record.file_type)

    with file_store.read_file(file_id, mode="b", use_tempfile=True) as xlsx_io:
        preview = parse_spreadsheet_for_preview(xlsx_io, "report.xlsx")

    assert [sheet.name for sheet in preview.sheets] == ["Revenue", "Notes"]
    assert all(not sheet.truncated for sheet in preview.sheets)

    revenue_rows = list(csv.reader(io.StringIO(preview.sheets[0].csv)))
    assert revenue_rows == [
        ["Region", "Quarter", "Amount"],
        ["EMEA", "Q1", "1250.5"],
        ['APAC, "South"', "Q2", "900"],
    ]

    notes_rows = list(csv.reader(io.StringIO(preview.sheets[1].csv)))
    assert notes_rows == [["Note"], ["multi\nline"]]


def test_parse_spreadsheet_for_preview_truncates_large_sheets() -> None:
    """Sheets bigger than the preview cap are cut at a row boundary and
    flagged as truncated."""
    workbook = openpyxl.Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Big"
    for i in range(50):
        sheet.append([f"row-{i}", "x" * 20])
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    with patch(
        "onyx.server.query_and_chat.chat_utils.MAX_PREVIEW_CHARS_PER_SHEET", 200
    ):
        preview = parse_spreadsheet_for_preview(buf, "big.xlsx")

    assert len(preview.sheets) == 1
    big_sheet = preview.sheets[0]
    assert big_sheet.truncated
    assert len(big_sheet.csv) <= 200
    # Still valid CSV, cut on a row boundary
    rows = list(csv.reader(io.StringIO(big_sheet.csv)))
    assert rows
    assert all(len(row) == 2 for row in rows)
    assert rows[0] == ["row-0", "x" * 20]

    # A first row larger than the cap yields empty CSV (never a mid-row slice)
    buf.seek(0)
    with patch("onyx.server.query_and_chat.chat_utils.MAX_PREVIEW_CHARS_PER_SHEET", 5):
        preview = parse_spreadsheet_for_preview(buf, "big.xlsx")
    assert preview.sheets[0].truncated
    assert preview.sheets[0].csv == ""


def test_truncation_skips_newlines_inside_quoted_cells() -> None:
    """A pathological quoted cell packed with newlines must not be treated as
    row boundaries — the cut lands after the last complete row before it."""
    workbook = openpyxl.Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Quoted"
    sheet.append(["col_a", "col_b"])
    # Second row's cell is a quoted multi-line value far larger than the cap
    sheet.append(["a", "line\n" * 500])
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    with patch("onyx.server.query_and_chat.chat_utils.MAX_PREVIEW_CHARS_PER_SHEET", 50):
        preview = parse_spreadsheet_for_preview(buf, "quoted.xlsx")

    quoted_sheet = preview.sheets[0]
    assert quoted_sheet.truncated
    # Only the complete header row survives; no partial quoted field leaks out
    rows = list(csv.reader(io.StringIO(quoted_sheet.csv)))
    assert rows == [["col_a", "col_b"]]


def test_xlsm_mime_type_classifies_as_tabular() -> None:
    """XLSM must classify TABULAR so chat routes it to the spreadsheet preview."""
    assert (
        mime_type_to_chat_file_type("application/vnd.ms-excel.sheet.macroEnabled.12")
        == ChatFileType.TABULAR
    )


def test_is_spreadsheet_mime_type() -> None:
    assert is_spreadsheet_mime_type(SPREADSHEET_MIME_TYPE)
    assert is_spreadsheet_mime_type(SPREADSHEET_MIME_TYPE + "; charset=utf-8")
    assert is_spreadsheet_mime_type("application/vnd.ms-excel.sheet.macroEnabled.12")
    # `parsed=true` must be a no-op for non-spreadsheet files
    assert not is_spreadsheet_mime_type("text/csv")
    assert not is_spreadsheet_mime_type("application/pdf")
    assert not is_spreadsheet_mime_type(None)
